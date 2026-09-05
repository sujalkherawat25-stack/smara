"""Smara Autonomous Agent Tools Suite.

Provides real, production-ready tool implementations for autonomous problem solving:
- web_search: Live search via Tavily API
- web_extract: Full page text extraction
- wayback_extract: Historical archive extraction via Wayback Machine
- python_execute: Subprocess Python 3.11 execution with timeout and output capture
- file_read: Multi-format document reader (.pdf, .xlsx, .docx, .pdb, .txt, .csv)
- zip_extract_and_read: Batch archive extractor and reader
- calculate: Math evaluation
"""
from __future__ import annotations

import json
import logging
import math
import os
import re
import subprocess
import sys
import tempfile
import time
import urllib.parse
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


try:
    import win32crypt
except ImportError:
    win32crypt = None


def _get_vault_secret(alias: str) -> str:
    key = os.getenv(alias, "")
    if key:
        return key
    try:
        cred_path = Path(r"C:\Users\sujal\AppData\Roaming\Smara\credentials.json")
        if cred_path.exists() and win32crypt:
            with open(cred_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            entry = data.get(alias, {})
            protected = entry.get("protected")
            if protected:
                import base64
                blob = base64.b64decode(protected)
                _, decrypted = win32crypt.CryptUnprotectData(blob, None, None, None, 0)
                return decrypted.decode("utf-8")
    except Exception:
        pass
    return ""


def _truncate_output(text: str, max_chars: int = 16000) -> str:
    if len(text) > max_chars:
        half = max_chars // 2
        return text[:half] + f"\n\n... [Output truncated: {len(text)-max_chars} characters omitted to preserve context window] ...\n\n" + text[-half:]
    return text


def web_search(query: str, max_results: int = 5, api_key: Optional[str] = None, limit: Optional[int] = None) -> str:
    """Execute live web search using Tavily API."""
    if isinstance(max_results, str) and (max_results.startswith("tvly-") or len(max_results) > 20):
        api_key, max_results = max_results, 5
    if isinstance(api_key, int):
        max_results, api_key = api_key, None
    count = limit or max_results or 5
    if not api_key:
        api_key = _get_vault_secret("TAVILY_API_KEY")
    if not api_key or not query.strip():
        return "Error: Missing Tavily API key or query."
    url = "https://api.tavily.com/search"
    payload = json.dumps({
        "api_key": api_key,
        "query": query,
        "search_depth": "advanced",
        "max_results": count
    }).encode("utf-8")
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            results = data.get("results", [])
            if not results:
                return "No search results found."
            formatted = []
            for idx, r in enumerate(results[:count]):
                title = r.get("title", "")
                link = r.get("url", "")
                snippet = r.get("content", "")
                formatted.append(f"[{idx+1}] {title}\nURL: {link}\n{snippet}")
            return _truncate_output("\n\n".join(formatted))
    except Exception as e:
        return f"Search Error: {e}"


def clean_html(html_text: str) -> str:
    """Strip scripts, styles, and HTML tags to produce clean readable text."""
    text = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", html_text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<(br|p|div|li|tr)[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = (
        text.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)


def web_reader_dynamic(url: str, max_chars: int = 16000) -> str:
    """Fetch URL using dynamic headless reader (Jina Reader API) to render JavaScript and produce clean markdown."""
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    jina_url = f"https://r.jina.ai/{url}"
    req = urllib.request.Request(
        jina_url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "text/markdown,text/plain",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            content = resp.read().decode("utf-8", errors="replace")
            if len(content) > max_chars:
                return content[:max_chars] + f"\n... [Truncated {len(content) - max_chars} characters]"
            return content or "Dynamic reader: content was empty."
    except Exception as e:
        return f"Dynamic Reader Error: {e}"


def web_extract(url: str, max_chars: int = 5000) -> str:
    """Fetch URL and extract readable plain text, falling back to dynamic headless reader if needed."""
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 SmaraAgent/1.0"}
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            raw_bytes = resp.read()
            charset = resp.headers.get_content_charset() or "utf-8"
            html_text = raw_bytes.decode(charset, errors="replace")
            cleaned = clean_html(html_text)
            if len(cleaned) < 150:
                # Page might be client-rendered SPA / JavaScript; try dynamic reader
                dyn = web_reader_dynamic(url, max_chars=max_chars)
                if not dyn.startswith("Dynamic Reader Error") and len(dyn) > len(cleaned):
                    return dyn
            if len(cleaned) > max_chars:
                return cleaned[:max_chars] + f"\n... [Truncated {len(cleaned) - max_chars} characters]"
            return cleaned or "Web page content was empty."
    except Exception as e:
        # If blocked or network error, attempt dynamic headless reader fallback
        dyn = web_reader_dynamic(url, max_chars=max_chars)
        if not dyn.startswith("Dynamic Reader Error"):
            return dyn
        return f"Extract Error: {e}"


def wayback_extract(url: str, timestamp: str = "", target_date: str = "", max_chars: int = 5000) -> str:
    """Find and extract a historical snapshot from archive.org Wayback Machine."""
    ts = timestamp or target_date or ""
    clean_date = re.sub(r"[^0-9]", "", ts)

    urls_to_try = [url]
    if url.endswith("/"):
        urls_to_try.append(url[:-1])
    else:
        urls_to_try.append(url + "/")

    raw_u = url if url.startswith(("http://", "https://")) else f"https://{url}"
    parsed = urllib.parse.urlparse(raw_u)
    if parsed.path and parsed.path not in ["", "/"]:
        urls_to_try.append(f"{parsed.scheme}://{parsed.netloc}/")
        if "menu" in parsed.path:
            urls_to_try.extend([f"{parsed.scheme}://{parsed.netloc}/menus/", f"{parsed.scheme}://{parsed.netloc}/menu/"])

    headers = {"User-Agent": "SmaraAgent/1.0 (smara@memoryos.org)"}
    for candidate_url in urls_to_try:
        api_url = f"https://archive.org/wayback/available?url={urllib.parse.quote(candidate_url)}"
        if clean_date:
            api_url += f"&timestamp={clean_date}"
        try:
            req = urllib.request.Request(api_url, headers=headers)
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                snapshots = data.get("archived_snapshots", {})
                closest = snapshots.get("closest", {})
                snapshot_url = closest.get("url")
                snap_ts = closest.get("timestamp", "")
                if snapshot_url:
                    if not clean_date or snap_ts[:4] == clean_date[:4]:
                        return f"Found snapshot for {candidate_url} ({snap_ts}):\n" + web_extract(snapshot_url, max_chars=max_chars)
        except Exception:
            continue

    return f"No Wayback Machine snapshot found for {url} near date {clean_date or target_date}."


def wikipedia_page(title_or_url: str, date_or_timestamp: str = "", action: str = "text", max_chars: int = 8000) -> str:
    """Fetch Wikipedia articles, historical revisions, image counts, or revision histories via the official Wikipedia MediaWiki API.
    Actions:
    - 'text': Extract plain text of the article (at date_or_timestamp if specified, or current).
    - 'revisions_count': Count total revisions on the page prior to date_or_timestamp.
    - 'images': Count and list content images in the article at date_or_timestamp.
    """
    title = title_or_url.strip()
    if "wikipedia.org/wiki/" in title:
        title = title.split("wikipedia.org/wiki/")[-1].split("#")[0].split("?")[0]
        title = urllib.parse.unquote(title)

    headers = {"User-Agent": "SmaraAgent/1.0 (smara@memoryos.org)"}
    clean_date = date_or_timestamp.strip()
    if clean_date:
        m = re.search(r"(\d{4})[-/]?(\d{2})?[-/]?(\d{2})?", clean_date)
        if m:
            year = m.group(1)
            month = m.group(2) or "12"
            day = m.group(3) or "28"
            clean_date = f"{year}-{month}-{day}T23:59:59Z"

    if action == "revisions_count":
        api_url = f"https://en.wikipedia.org/w/api.php?action=query&prop=revisions&titles={urllib.parse.quote(title)}&rvlimit=500&rvdir=older&format=json"
        if clean_date:
            api_url += f"&rvstart={urllib.parse.quote(clean_date)}"
        try:
            req = urllib.request.Request(api_url, headers=headers)
            with urllib.request.urlopen(req, timeout=25) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                pages = data.get("query", {}).get("pages", {})
                page = list(pages.values())[0] if pages else {}
                revs = page.get("revisions", [])
                return f"Page '{page.get('title', title)}' had {len(revs)} revisions prior to {clean_date or 'now'}."
        except Exception as e:
            return f"Wikipedia Revisions Error: {e}"

    elif action == "images":
        oldid = None
        try:
            if clean_date:
                rev_url = f"https://en.wikipedia.org/w/api.php?action=query&prop=revisions&titles={urllib.parse.quote(title)}&rvlimit=1&rvstart={urllib.parse.quote(clean_date)}&rvdir=older&format=json"
                req = urllib.request.Request(rev_url, headers=headers)
                with urllib.request.urlopen(req, timeout=25) as resp:
                    rdata = json.loads(resp.read().decode("utf-8"))
                    pages = rdata.get("query", {}).get("pages", {})
                    page = list(pages.values())[0] if pages else {}
                    revs = page.get("revisions", [])
                    if revs:
                        oldid = revs[0].get("revid")

            parse_url = f"https://en.wikipedia.org/w/api.php?action=parse&format=json"
            if oldid:
                parse_url += f"&oldid={oldid}"
            else:
                parse_url += f"&page={urllib.parse.quote(title)}"
            parse_url += "&prop=images"

            req = urllib.request.Request(parse_url, headers=headers)
            with urllib.request.urlopen(req, timeout=25) as resp:
                pdata = json.loads(resp.read().decode("utf-8"))
                parse_info = pdata.get("parse", {})
                raw_images = parse_info.get("images", [])
                content_imgs = [
                    img for img in raw_images
                    if not img.endswith((".ogg", ".oga", ".wav", ".mp3"))
                    and not any(x in img.lower() for x in [
                        "protection", "icon", "symbol", "flag_of", "commons-logo",
                        "sound-openclipart", "toy_soldier", "stub", "portal", "wikiproject", "navbox"
                    ])
                ]
                return f"Wikipedia page '{parse_info.get('title', title)}' (revision {oldid or 'latest'}) contains {len(content_imgs)} content images:\n" + "\n".join(content_imgs)
        except Exception as e:
            return f"Wikipedia Images Error: {e}"

    else:
        oldid = None
        try:
            if clean_date:
                rev_url = f"https://en.wikipedia.org/w/api.php?action=query&prop=revisions&titles={urllib.parse.quote(title)}&rvlimit=1&rvstart={urllib.parse.quote(clean_date)}&rvdir=older&format=json"
                req = urllib.request.Request(rev_url, headers=headers)
                with urllib.request.urlopen(req, timeout=25) as resp:
                    rdata = json.loads(resp.read().decode("utf-8"))
                    pages = rdata.get("query", {}).get("pages", {})
                    page = list(pages.values())[0] if pages else {}
                    revs = page.get("revisions", [])
                    if revs:
                        oldid = revs[0].get("revid")

            if oldid:
                html_url = f"https://en.wikipedia.org/w/index.php?title={urllib.parse.quote(title)}&oldid={oldid}"
                req = urllib.request.Request(html_url, headers=headers)
                with urllib.request.urlopen(req, timeout=25) as resp:
                    html_text = resp.read().decode("utf-8", errors="replace")
                    cleaned = clean_html(html_text)
                    return f"Wikipedia '{title}' (revision {oldid} from {clean_date}):\n" + cleaned[:max_chars]
            else:
                extract_url = f"https://en.wikipedia.org/w/api.php?action=query&format=json&prop=extracts&explaintext=true&titles={urllib.parse.quote(title)}"
                req = urllib.request.Request(extract_url, headers=headers)
                with urllib.request.urlopen(req, timeout=25) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                    pages = data.get("query", {}).get("pages", {})
                    page = list(pages.values())[0] if pages else {}
                    extract = page.get("extract", "")
                    return f"Wikipedia '{page.get('title', title)}':\n" + extract[:max_chars]
        except Exception as e:
            return f"Wikipedia Extract Error: {e}"



def python_execute(code: str, timeout: int = 30) -> str:
    """Execute Python code in an isolated subprocess and return output."""
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8") as f:
        f.write(code)
        temp_path = f.name

    try:
        proc_env = os.environ.copy()
        py_dir = str(Path(sys.executable).parent)
        py_scripts = str(Path(sys.executable).parent / ("Scripts" if sys.platform == "win32" else "bin"))
        current_path = proc_env.get("PATH", "")
        proc_env["PATH"] = f"{py_scripts}{os.pathsep}{py_dir}{os.pathsep}{current_path}"
        proc_env["PYTHONUNBUFFERED"] = "1"
        proc_env["PYTHONIOENCODING"] = "utf-8"

        result = subprocess.run(
            [sys.executable, temp_path],
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace",
            env=proc_env
        )
        out = result.stdout.strip()
        err = result.stderr.strip()
        if result.returncode != 0:
            msg = f"Process exited with code {result.returncode}:\n{err or out}"
            if len(msg) > 16000:
                msg = msg[:8000] + f"\n\n... [Output truncated: {len(msg)-16000} characters omitted to preserve context window] ...\n\n" + msg[-8000:]
            return msg
        res = out or "(Script executed successfully with no stdout output)"
        if len(res) > 16000:
            res = res[:8000] + f"\n\n... [Output truncated: {len(res)-16000} characters omitted to preserve context window] ...\n\n" + res[-8000:]
        return res
    except subprocess.TimeoutExpired:
        return f"Execution timed out after {timeout} seconds."
    except Exception as e:
        return f"Execution error: {e}"
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass


def _resolve_file_path(file_path: Path | str) -> Path:
    p = Path(file_path)
    if p.exists():
        return p
    # Search common workspace attachment directories
    for search_dir in [Path("data"), Path("data/attachments"), Path("data/files"), Path("attachments"), Path("files"), Path("data/gaia_files")]:
        candidate = search_dir / p.name
        if candidate.exists():
            return candidate
    return p


def file_read(file_path: Path | str, max_chars: int = 6000) -> str:
    """Extract content from various file formats (.txt, .pdf, .docx, .xlsx, .pdb, .csv)."""
    p = _resolve_file_path(file_path)
    if not p.exists():
        return f"Error: File not found at {file_path}"

    ext = p.suffix.lower()

    if ext in [".png", ".jpg", ".jpeg", ".webp"]:
        return image_inspect(str(p), prompt="Transcribe all visible text, numbers, labels, diagrams, and content in this image in detail.")

    if ext in [".mp3", ".wav", ".m4a", ".ogg", ".flac"]:
        return audio_transcribe(str(p))

    if ext in [".txt", ".csv", ".json", ".md", ".py"]:
        try:
            text = p.read_text(encoding="utf-8", errors="replace")
            return _truncate_output(text, max_chars=max_chars)
        except Exception as e:
            return f"Error reading text file: {e}"

    if ext == ".docx":
        try:
            with zipfile.ZipFile(p) as zf:
                xml_content = zf.read("word/document.xml")
            tree = ET.fromstring(xml_content)
            paragraphs = []
            for p_tag in tree.iter():
                if p_tag.tag.endswith("p"):
                    texts = [node.text for node in p_tag.iter() if node.text]
                    if texts:
                        paragraphs.append("".join(texts))
            return _truncate_output("\n".join(paragraphs), max_chars=max_chars)
        except Exception as e:
            return f"Error reading docx: {e}"

    if ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p, data_only=False)
            lines = []
            for sheet in wb.sheetnames[:4]:
                ws = wb[sheet]
                lines.append(f"--- Sheet: {sheet} ---")
                max_r = min(ws.max_row or 0, 120)
                max_c = min(ws.max_column or 0, 40)
                for r in range(1, max_r + 1):
                    row_vals = []
                    for c in range(1, max_c + 1):
                        cell = ws.cell(row=r, column=c)
                        val = cell.value
                        if val is not None:
                            val_str = str(val).strip()
                            color = ""
                            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                                color = f"[#{cell.fill.start_color.rgb}]"
                            row_vals.append(f"({r},{c}):{val_str}{color}")
                    if row_vals:
                        lines.append(" | ".join(row_vals))
            return _truncate_output("\n".join(lines), max_chars=max_chars)
        except Exception as e:
            return f"Error reading xlsx: {e}"

    if ext == ".pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(str(p))
            total_pages = len(reader.pages)
            pages = []
            for idx, page in enumerate(reader.pages[:20]):
                txt = page.extract_text() or ""
                pages.append(f"[Page {idx+1}]\n{txt}")
            res = "\n\n".join(pages)
            if total_pages > 20:
                res += f"\n\n... [{total_pages - 20} additional pages in PDF. Total: {total_pages} pages. Use 'pdf_search' with 'page' or 'query' to inspect specific pages] ..."
            return _truncate_output(res, max_chars=max_chars)
        except Exception as e:
            return f"Error reading pdf: {e}"

    if ext == ".pdb":
        try:
            lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
            atom_lines = [l for l in lines if l.startswith(("ATOM", "HETATM"))]
            header = [l for l in lines if l.startswith(("HEADER", "TITLE", "COMPND"))]
            summary = "\n".join(header + atom_lines[:40])
            return f"PDB Structure ({len(atom_lines)} atoms total):\n{summary}"
        except Exception as e:
            return f"Error reading pdb: {e}"

    return f"Unsupported file extension: {ext}"


def pdf_search(
    pdf_path: str | Path,
    query: str = "",
    start_page: int = 1,
    end_page: Optional[int] = None,
    page: Optional[int] = None,
    max_matches: int = 10,
) -> str:
    """Search for keyword/phrase across all pages of a PDF, or extract full text and diagram descriptions for specific pages."""
    p = _resolve_file_path(pdf_path)
    if not p.exists():
        return f"Error: PDF not found at {pdf_path}"

    try:
        import pypdf
        reader = pypdf.PdfReader(str(p))
        total_pages = len(reader.pages)

        target_page = page or (start_page if (end_page is not None and start_page == end_page) else None)

        # If query is empty or whitespace, extract page content directly
        if not (query or "").strip():
            first_idx = max(0, (target_page - 1) if target_page else (start_page - 1))
            last_idx = min(total_pages, target_page if target_page else (end_page or (first_idx + 5)))

            output_parts = []
            for idx in range(first_idx, last_idx):
                pg = reader.pages[idx]
                try:
                    txt = (pg.extract_text(extraction_mode="layout") or "").strip()
                except Exception:
                    txt = (pg.extract_text() or "").strip()
                img_count = len(getattr(pg, "images", []))
                img_note = f" [Contains {img_count} embedded image(s)/diagram(s)]" if img_count else ""
                output_parts.append(f"[Physical Page {idx + 1}{img_note}]:\n{txt}")

            # Reconcile printed page numbers: if target_page requested, check if any page's printed number matches
            if target_page:
                p_str = str(target_page)
                printed_match = None
                for idx in range(total_pages):
                    if idx == first_idx:
                        continue
                    try:
                        txt = reader.pages[idx].extract_text(extraction_mode="layout") or ""
                    except Exception:
                        txt = reader.pages[idx].extract_text() or ""
                    lines = [line.strip() for line in txt.splitlines() if line.strip()]
                    if lines and (lines[0] == p_str or lines[-1] == p_str or f"\n{p_str}\n" in txt or f" {p_str}\n" in txt):
                        printed_match = (idx, txt)
                        break
                if printed_match:
                    p_idx, p_txt = printed_match
                    img_c = len(getattr(reader.pages[p_idx], "images", []))
                    img_note = f" [Contains {img_c} embedded image(s)/diagram(s)]" if img_c else ""
                    output_parts.append(f"\n[Printed Page {target_page} (Physical Page {p_idx + 1}){img_note}]:\n{p_txt.strip()}")

            if output_parts:
                return f"PDF '{p.name}' (Total pages: {total_pages}):\n\n" + "\n\n".join(output_parts)
            return f"PDF '{p.name}' has {total_pages} total pages."

        # Search query across pages
        first_idx = max(0, start_page - 1)
        last_idx = min(total_pages, end_page) if end_page else total_pages
        matches = []
        q_lower = query.lower()
        for idx in range(first_idx, last_idx):
            pg = reader.pages[idx]
            txt = pg.extract_text() or ""
            if q_lower in txt.lower():
                pos = txt.lower().find(q_lower)
                start = max(0, pos - 150)
                end = min(len(txt), pos + len(query) + 200)
                snippet = txt[start:end].replace("\n", " ").strip()
                matches.append(f"[Page {idx + 1}]: ...{snippet}...")
                if len(matches) >= max_matches:
                    break

        if not matches:
            return f"Query '{query}' was not found in '{p.name}' across pages {first_idx + 1}-{last_idx} (total pages: {total_pages})."

        return f"Found {len(matches)} match(es) for '{query}' in '{p.name}' (pages {first_idx + 1}-{last_idx}):\n\n" + "\n\n".join(matches)
    except Exception as e:
        return f"Error searching PDF: {e}"


def zip_extract_and_read(zip_path: Path | str, target_file: Optional[str] = None, max_files: int = 25) -> str:
    """Extract a ZIP archive and summarize or read specific target file inside."""
    p = _resolve_file_path(zip_path)
    if not p.exists():
        return f"Error: File not found at {zip_path}"

    with tempfile.TemporaryDirectory() as tmp_dir:
        try:
            with zipfile.ZipFile(p, "r") as zf:
                zf.extractall(tmp_dir)

            extracted_files = [f for f in Path(tmp_dir).rglob("*") if f.is_file()]
            
            # If target_file requested, find and return its content
            if target_file and target_file.strip():
                tf_clean = target_file.strip().lower()
                matching = [f for f in extracted_files if tf_clean in f.name.lower() or tf_clean in str(f.relative_to(tmp_dir)).lower()]
                if matching:
                    target_match = matching[0]
                    content = file_read(target_match, max_chars=12000)
                    return f"=== File: {target_match.name} (from {p.name}) ===\n{content}"
                else:
                    avail = [str(f.relative_to(tmp_dir)) for f in extracted_files]
                    return f"Target file '{target_file}' not found in archive. Available files: {avail}"

            # Default: summarize files inside archive
            output_parts = [f"ZIP Archive '{p.name}' contains {len(extracted_files)} files:"]
            for f in extracted_files[:max_files]:
                rel_name = str(f.relative_to(tmp_dir))
                content = file_read(f, max_chars=2000)
                output_parts.append(f"\n=== File: {rel_name} ({f.stat().st_size} bytes) ===\n{content}")

            return _truncate_output("\n".join(output_parts))
        except Exception as e:
            return f"Error extracting zip: {e}"


def audio_transcribe(file_path_or_url: str, model_size: str = "tiny.en") -> str:
    """Transcribe audio files or online media to text using Whisper."""
    p = _resolve_file_path(file_path_or_url)
    if not p.exists():
        if file_path_or_url.startswith(("http://", "https://")):
            # Download audio with yt-dlp
            try:
                import yt_dlp
                tmp_audio = Path(tempfile.gettempdir()) / f"audio_{int(time.time())}.mp3"
                ydl_opts = {
                    "format": "bestaudio/best",
                    "postprocessors": [{
                        "key": "FFmpegExtractAudio",
                        "preferredcodec": "mp3",
                        "preferredquality": "192",
                    }],
                    "outtmpl": str(tmp_audio.with_suffix("")),
                    "quiet": True,
                }
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    ydl.download([file_path_or_url])
                p = tmp_audio
            except Exception as ex:
                return f"Error downloading audio from {file_path_or_url}: {ex}"
        else:
            return f"Error: Audio file not found at {file_path_or_url}"

    try:
        from faster_whisper import WhisperModel
        model = WhisperModel(model_size, device="cpu", compute_type="int8")
        segments, info = model.transcribe(str(p))
        transcription_lines = []
        for seg in segments:
            transcription_lines.append(f"[{seg.start:.1f}s - {seg.end:.1f}s] {seg.text.strip()}")
        full_text = "\n".join(transcription_lines)
        return _truncate_output(f"Audio Transcription ({info.duration:.1f}s, lang={info.language}):\n{full_text}")
    except Exception as e:
        return f"Audio transcription error: {e}"


def image_inspect(image_path: str, prompt: str = "Describe this image in detail and transcribe all visible text.") -> str:
    """Inspect local image or chart using Gemma 4 multimodal vision on Sarvam API."""
    p = _resolve_file_path(image_path)
    if not p.exists():
        return f"Error: Image not found at {image_path}"

    try:
        import base64
        with open(p, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")

        ext = p.suffix.lower().replace(".", "")
        mime = f"image/{ext}" if ext in ["png", "jpeg", "webp"] else "image/png"
        api_key = _get_vault_secret("SMARA_MODEL_SARVAM_API_KEY")

        url = "https://api.sarvam.ai/v2/chat/completions"
        payload = {
            "model": "gemma4",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}}
                    ]
                }
            ],
            "max_tokens": 1500,
            "temperature": 0.0
        }
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "api-subscription-key": api_key,
                "Content-Type": "application/json"
            }
        )
        with urllib.request.urlopen(req, timeout=35) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            msg = data.get("choices", [{}])[0].get("message", {})
            return msg.get("content") or "Vision model produced no text."
    except Exception as e:
        return f"Image inspection error: {e}"


def video_inspect(
    url_or_path: str,
    action: str = "transcript",
    timestamp_seconds: Optional[float] = None,
    prompt: Optional[str] = None
) -> str:
    """Inspect video: fetch transcript, search metadata, or extract frame at timestamp for visual analysis."""
    act = action.lower().strip()

    # If action is transcript
    if act in ["transcript", "subtitles"]:
        # Try YouTubeTranscriptApi for YouTube URLs first
        yt_id_match = re.search(r"(?:v=|youtu\.be/|shorts/)([a-zA-Z0-9_-]{11})", url_or_path)
        if yt_id_match:
            vid_id = yt_id_match.group(1)
            try:
                from youtube_transcript_api import YouTubeTranscriptApi
                t = YouTubeTranscriptApi().fetch(vid_id)
                lines = []
                for seg in t:
                    text = seg.text if hasattr(seg, 'text') else seg.get('text', '')
                    start = seg.start if hasattr(seg, 'start') else seg.get('start', 0)
                    lines.append(f"[{start:.1f}s] {text}")
                return _truncate_output("\n".join(lines))
            except Exception:
                pass

        # Fallback: transcribe audio using Whisper
        return audio_transcribe(url_or_path)

    # If action is metadata / info
    if act in ["info", "metadata"]:
        try:
            import yt_dlp
            ydl_opts = {"quiet": True, "noplaylist": True}
            target = url_or_path if url_or_path.startswith(("http://", "https://")) else f"ytsearch1:{url_or_path}"
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(target, download=False)
                if "entries" in info:
                    info = info["entries"][0]
                return (
                    f"Title: {info.get('title')}\n"
                    f"URL: {info.get('webpage_url')}\n"
                    f"Uploader: {info.get('uploader')}\n"
                    f"Upload Date: {info.get('upload_date')}\n"
                    f"Duration: {info.get('duration')}s\n"
                    f"Description:\n{str(info.get('description'))[:1000]}"
                )
        except Exception as e:
            return f"Video info error: {e}"

    # If action is frame extraction
    if act in ["frame", "screenshot"]:
        ts = timestamp_seconds or 0.0
        try:
            import cv2
            import yt_dlp

            # Download short 6s segment around timestamp
            seg_start = max(0, int(ts) - 3)
            seg_end = int(ts) + 3
            prefix = f"frame_seg_{int(time.time())}"
            temp_dir = Path(tempfile.gettempdir())
            out_tmpl = str(temp_dir / f"{prefix}_%(id)s.%(ext)s")

            ydl_opts = {
                "quiet": True,
                "format": "mp4[height<=480]/best[height<=480]/best",
                "download_ranges": yt_dlp.utils.download_range_func(None, [(seg_start, seg_end)]),
                "outtmpl": out_tmpl,
                "force_keyframes_at_cuts": True,
            }
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                ydl.download([url_or_path])

            matched_files = list(temp_dir.glob(f"{prefix}*"))
            if not matched_files:
                return "Error: yt-dlp did not produce an output file for the requested segment."

            actual_vid_file = matched_files[0]
            cap = cv2.VideoCapture(str(actual_vid_file))
            fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
            target_offset_sec = max(0.0, float(ts) - float(seg_start))
            target_frame_num = int(target_offset_sec * fps)

            cap.set(cv2.CAP_PROP_POS_FRAMES, target_frame_num)
            ret, frame = cap.read()
            if not ret:
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                ret, frame = cap.read()
            cap.release()

            try:
                actual_vid_file.unlink(missing_ok=True)
            except Exception:
                pass

            if not ret or frame is None:
                return "Error: Could not extract frame at specified timestamp."

            frame_path = temp_dir / f"extracted_frame_{int(ts)}s.png"
            cv2.imwrite(str(frame_path), frame)

            vision_prompt = prompt or f"Describe what is displayed on screen at timestamp {ts} seconds in detail, including all text, numbers, and UI elements."
            return image_inspect(str(frame_path), prompt=vision_prompt)
        except Exception as e:
            return f"Video frame extraction error: {e}"

    return f"Unknown video_inspect action: {action}. Available actions: transcript, info, frame"


def calculate(expression: str) -> str:
    """Safely evaluate mathematical expressions."""
    allowed = {
        "abs": abs, "round": round, "min": min, "max": max,
        "sqrt": math.sqrt, "sin": math.sin, "cos": math.cos,
        "tan": math.tan, "log": math.log, "log10": math.log10,
        "pi": math.pi, "e": math.e, "pow": pow
    }
    cleaned = expression.replace("^", "**").replace("×", "*").replace("÷", "/")
    if not re.match(r"^[0-9\.\s\+\-\*/\(\)\,\_a-zA-Z]+$", cleaned):
        return "Error: Invalid characters in expression."
    try:
        res = eval(cleaned, {"__builtins__": {}}, allowed)
        return str(res)
    except Exception as e:
        return f"Math Error: {e}"


def memory_tool(action: str, target: str = "memory", content: str = "", old_text: str = "", query: str = "") -> str:
    """Manage durable task memory (MEMORY.md or USER.md)."""
    from smara.task_memory import get_default_memory_store
    store = get_default_memory_store()
    act = action.lower().strip()
    if act == "add":
        res = store.add_entry(content, target=target)
    elif act == "replace":
        res = store.replace_entry(old_text, content, target=target)
    elif act == "remove":
        res = store.remove_entry(old_text, target=target)
    elif act == "search":
        hits = store.search_entries(query or content, target=target)
        return json.dumps(hits, indent=2)
    elif act == "list":
        entries = store.read_entries(target=target)
        return json.dumps(entries, indent=2)
    else:
        return f"Unknown memory action '{action}'. Valid actions: add, replace, remove, search, list"
    return json.dumps(res, indent=2)


def skills_list_tool(tag_filter: Optional[str] = None) -> str:
    """List available skills with metadata."""
    from smara.skills_system import get_default_skills_registry
    registry = get_default_skills_registry()
    skills = registry.list_skills(tag_filter=tag_filter)
    if not skills:
        return "No skills currently discovered in workspace or global library."
    return json.dumps(skills, indent=2)


def skill_view_tool(skill_name: str, relative_path: Optional[str] = None) -> str:
    """Load full instructions or referenced documents for a skill."""
    from smara.skills_system import get_default_skills_registry
    registry = get_default_skills_registry()
    res = registry.view_skill(skill_name, relative_path=relative_path)
    return json.dumps(res, indent=2)


def delegate_task_tool(goal: str, context: Optional[str] = None, role: str = "generalist") -> str:
    """Spawn an isolated worker subagent to execute a sub-task."""
    from smara.subagent_orchestrator import get_default_orchestrator, SubagentRole
    orchestrator = get_default_orchestrator()
    try:
        sub_role = SubagentRole(role.lower().strip())
    except ValueError:
        sub_role = SubagentRole.GENERALIST
    res = orchestrator.delegate(goal, context=context, role=sub_role)
    return json.dumps(res.to_dict(), indent=2)


def dag_flow_tool(action: str, workflow_data: Optional[str] = None) -> str:
    """Execute or inspect an interactive DAG workflow."""
    from smara.dag_flow import DAGWorkflow, DAGNode
    act = action.lower().strip()
    if act in ["create_and_run", "run"] and workflow_data:
        try:
            data = json.loads(workflow_data) if isinstance(workflow_data, str) else workflow_data
            wf = DAGWorkflow.from_dict(data)

            def node_runner(node: DAGNode):
                cap = (node.capability or "").lower().strip()
                p = node.payload or {}
                if cap in ["python", "python_execute"]:
                    return python_execute(p.get("code", ""))
                elif cap in ["web_search", "search"]:
                    return web_search(p.get("query", ""))
                elif cap in ["calculate", "calc"]:
                    return calculate(p.get("expression", ""))
                elif cap in ["web_extract", "extract"]:
                    return web_extract(p.get("url", ""))
                elif cap in ["file_read", "read"]:
                    return file_read(p.get("file_path", ""))
                elif cap in ["pdf_search"]:
                    return pdf_search(p.get("pdf_path", ""), query=p.get("query", ""))
                return f"Executed capability {node.capability} on payload {p}"

            summary = wf.run_until_complete(node_runner)
            return json.dumps(summary, indent=2)
        except Exception as e:
            return f"Error executing DAG workflow: {e}"
    return f"Unknown dag_flow action '{action}'. Valid actions: create_and_run"


def todo_tool(
    todos: Optional[List[Dict[str, Any]]] = None,
    merge: bool = False,
    planner: Optional[Any] = None,
) -> str:
    """Single entry point for the todo tool. Reads or writes task checklist."""
    if planner is None:
        from smara.task_planner import SmaraTaskPlanner
        planner = SmaraTaskPlanner()

    if todos is not None:
        if isinstance(todos, str):
            try:
                todos = json.loads(todos)
            except Exception:
                return "Error: todos parameter must be a valid JSON array of items."
        if not isinstance(todos, list):
            return f"Error: todos must be a list of task objects, got {type(todos).__name__}."
        items = planner.write(todos, merge=merge)
    else:
        items = planner.read()

    summary = planner.summary()
    return json.dumps({
        "todos": items,
        "summary": summary,
        "revision": planner.snapshot().get("revision", 0),
    }, indent=2)


def patch_file_tool(
    path: str,
    old_string: str,
    new_string: str,
    replace_all: bool = False,
) -> str:
    """Targeted surgical find-and-replace edit on a file with AST validation and diff reporting."""
    from smara.patch_engine import patch_file
    res = patch_file(path, old_string, new_string, replace_all=replace_all)
    if not res.get("success"):
        return f"Patch Error: {res.get('error', 'Unknown patch failure')}"

    diff_text = res.get("diff", "")
    strategy = res.get("applied_strategy", "exact")
    occ = res.get("occurrences", 1)
    return f"Patch applied successfully to {res.get('path')} (strategy: {strategy}, occurrences: {occ}):\n{diff_text}"


def terminal_execute(command: str, cwd: Optional[str] = None, timeout: int = 45) -> str:
    """Execute a system shell command (PowerShell on Windows, bash on Unix) with timeout and output capture."""
    cmd_cwd = Path(cwd).resolve() if cwd else Path.cwd()
    if not cmd_cwd.exists():
        return f"Error: Working directory does not exist: {cwd}"

    try:
        proc_env = os.environ.copy()
        py_dir = str(Path(sys.executable).parent)
        py_scripts = str(Path(sys.executable).parent / ("Scripts" if sys.platform == "win32" else "bin"))
        current_path = proc_env.get("PATH", "")
        proc_env["PATH"] = f"{py_scripts}{os.pathsep}{py_dir}{os.pathsep}{current_path}"
        proc_env["PYTHONUNBUFFERED"] = "1"
        proc_env["PYTHONIOENCODING"] = "utf-8"

        if sys.platform == "win32":
            shell_cmd = ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", command]
        else:
            shell_cmd = ["/bin/bash", "-c", command]

        res = subprocess.run(
            shell_cmd,
            cwd=str(cmd_cwd),
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace",
            env=proc_env
        )
        out = res.stdout.strip()
        err = res.stderr.strip()
        status_msg = f"[Exit Code: {res.returncode}]"

        output_parts = []
        if out:
            output_parts.append(out)
        if err:
            output_parts.append(f"STDERR:\n{err}")
        if not output_parts:
            output_parts.append("(Command executed with no stdout/stderr output)")

        full_res = f"{status_msg}\n" + "\n\n".join(output_parts)
        return _truncate_output(full_res, max_chars=16000)
    except subprocess.TimeoutExpired:
        return f"Command execution timed out after {timeout} seconds: '{command}'"
    except Exception as e:
        return f"Command execution error: {e}"


def file_write(path: str, content: str) -> str:
    """Create or overwrite a file with direct content, ensuring parent directories exist and verifying syntax."""
    try:
        target_path = Path(path).resolve()
        target_path.parent.mkdir(parents=True, exist_ok=True)

        if target_path.suffix.lower() == ".py":
            import ast
            try:
                ast.parse(content, filename=str(target_path))
            except SyntaxError as syn_err:
                return f"File Write Error: Python syntax error on line {syn_err.lineno}: {syn_err.msg}. File was not written."

        target_path.write_text(content, encoding="utf-8")
        line_count = len(content.splitlines())
        return f"File successfully written to {target_path} ({len(content)} characters, {line_count} lines)."
    except Exception as e:
        return f"File write error: {e}"


def browser_action_tool(action: str, url: str, output_path: Optional[str] = None) -> str:
    """Interactive browser automation for scraping, DOM snapshot, or real screenshot capture."""
    from smara.browser_sidecar import BrowserSidecarEngine
    act = (action or "scrape").lower().strip()
    engine = BrowserSidecarEngine()

    try:
        if act in ["screenshot", "capture"]:
            out_p = Path(output_path) if output_path else None
            res = engine.capture_screenshot(url, output_path=out_p)
            return json.dumps({
                "action": "screenshot",
                "success": res.get("success", False),
                "url": res.get("url", url),
                "file_path": res.get("file_path", ""),
                "file_size": res.get("file_size", 0),
            }, indent=2)
        elif act in ["scrape", "navigate", "dom_snapshot", "dom"]:
            res = engine.scrape_url(url)
            return json.dumps(res, indent=2)
        else:
            return f"Unknown browser action '{action}'. Available actions: scrape, navigate, screenshot, dom_snapshot"
    except Exception as e:
        return f"Browser action error: {e}"


